"""Tests for PR5: Export endpoints — CSV, XLSX, full workbook, snapshot."""

import os
import tempfile
from datetime import date, timedelta

import openpyxl

from app.models import Account, Category, Entity, Transaction, TransactionSplit


def _seed(db):
    house = Entity(name="House", entity_type="household", is_default=True)
    airbnb = Entity(name="Airbnb", entity_type="rental", is_default=False)
    db.add_all([house, airbnb])
    db.flush()

    acct = Account(name="Chequing", account_type="checking", current_balance=5000)
    db.add(acct)
    db.flush()

    groceries = Category(name="Groceries", is_system=True)
    utilities = Category(name="Utilities", is_system=True)
    db.add_all([groceries, utilities])
    db.flush()

    return house, airbnb, acct, groceries, utilities


def _txn(db, acct, entity, cat, amount, name, days_ago=0):
    d = date.today() - timedelta(days=days_ago)
    txn = Transaction(
        account_id=acct.id,
        entity_id=entity.id if entity else None,
        entity_source="manual",
        category_id=cat.id if cat else None,
        amount=amount,
        date=d,
        name=name,
        txn_type="income" if amount < 0 else "expense",
        review_status="confirmed",
        review_source="manual",
        dedup_hash=Transaction.compute_dedup_hash(d, amount, name, acct.id),
    )
    db.add(txn)
    db.flush()
    return txn


# ---- CSV export ----

def test_csv_export_basic(client, db_session):
    house, airbnb, acct, groceries, utilities = _seed(db_session)
    _txn(db_session, acct, house, groceries, 100, "Food", days_ago=1)
    _txn(db_session, acct, airbnb, utilities, 50, "Hydro", days_ago=1)
    db_session.commit()

    resp = client.get("/api/export/transactions/csv")
    assert resp.status_code == 200
    assert "text/csv" in resp.headers["content-type"]
    lines = resp.text.strip().split("\n")
    assert len(lines) == 3  # header + 2 rows
    assert "Date" in lines[0]
    assert "Food" in resp.text


def test_csv_export_with_entity_filter(client, db_session):
    house, airbnb, acct, groceries, utilities = _seed(db_session)
    _txn(db_session, acct, house, groceries, 100, "House Food", days_ago=1)
    _txn(db_session, acct, airbnb, utilities, 50, "Airbnb Hydro", days_ago=1)
    db_session.commit()

    resp = client.get(f"/api/export/transactions/csv?entity_id={airbnb.id}")
    assert resp.status_code == 200
    lines = resp.text.strip().split("\n")
    assert len(lines) == 2  # header + 1 row
    assert "Airbnb Hydro" in resp.text
    assert "House Food" not in resp.text


# ---- XLSX export ----

def test_xlsx_export_basic(client, db_session):
    house, airbnb, acct, groceries, utilities = _seed(db_session)
    _txn(db_session, acct, house, groceries, 100, "Food", days_ago=1)
    _txn(db_session, acct, house, utilities, 200, "Hydro", days_ago=1)
    db_session.commit()

    resp = client.get("/api/export/transactions/xlsx")
    assert resp.status_code == 200
    assert "spreadsheetml" in resp.headers["content-type"]

    # Parse the workbook
    import io
    wb = openpyxl.load_workbook(io.BytesIO(resp.content))
    ws = wb["Transactions"]
    assert ws.cell(row=1, column=1).value == "Date"
    assert ws.max_row == 3  # header + 2 rows


# ---- Full workbook ----

def test_full_workbook_sheets(client, db_session):
    house, airbnb, acct, groceries, utilities = _seed(db_session)
    _txn(db_session, acct, house, groceries, 100, "Food", days_ago=1)
    db_session.commit()

    resp = client.get("/api/export/full")
    assert resp.status_code == 200

    import io
    wb = openpyxl.load_workbook(io.BytesIO(resp.content))
    sheet_names = wb.sheetnames
    assert "Info" in sheet_names
    assert "Transactions" in sheet_names
    assert "Entities" in sheet_names
    assert "Categories" in sheet_names
    assert "Accounts" in sheet_names
    # Per-entity pivot sheets
    pivot_sheets = [s for s in sheet_names if s.startswith("Pivot")]
    assert len(pivot_sheets) == 2  # House + Airbnb


def test_full_workbook_row_counts(client, db_session):
    house, airbnb, acct, groceries, utilities = _seed(db_session)
    _txn(db_session, acct, house, groceries, 100, "A", days_ago=1)
    _txn(db_session, acct, house, utilities, 200, "B", days_ago=2)
    _txn(db_session, acct, airbnb, groceries, 50, "C", days_ago=3)
    db_session.commit()

    resp = client.get("/api/export/full")
    import io
    wb = openpyxl.load_workbook(io.BytesIO(resp.content))

    # Transactions sheet: header + 3 rows
    assert wb["Transactions"].max_row == 4

    # Entities: header + 2
    assert wb["Entities"].max_row == 3

    # Categories: header + 2
    assert wb["Categories"].max_row == 3


# ---- Snapshot ----

def test_snapshot_generation(db_session):
    house = Entity(name="House", entity_type="household", is_default=True)
    db_session.add(house)
    db_session.flush()

    acct = Account(name="Checking", account_type="checking", current_balance=1000)
    db_session.add(acct)
    db_session.flush()

    d = date.today()
    txn = Transaction(
        account_id=acct.id, entity_id=house.id, entity_source="manual",
        amount=100, date=d, name="Test", txn_type="expense",
        review_status="confirmed", review_source="manual",
        dedup_hash=Transaction.compute_dedup_hash(d, 100, "Test", acct.id),
    )
    db_session.add(txn)
    db_session.commit()

    from app.routers.export import generate_snapshot
    with tempfile.TemporaryDirectory() as tmpdir:
        path = generate_snapshot(db_session, output_dir=tmpdir)
        assert os.path.exists(path)
        assert path.endswith(".xlsx")

        wb = openpyxl.load_workbook(path)
        assert "Info" in wb.sheetnames
        assert "Transactions" in wb.sheetnames


def test_snapshot_retention(db_session):
    """Snapshot keeps only the last 12 files."""
    house = Entity(name="House", entity_type="household", is_default=True)
    db_session.add(house)
    db_session.commit()

    from app.routers.export import generate_snapshot
    with tempfile.TemporaryDirectory() as tmpdir:
        # Create 14 fake snapshot files
        for i in range(14):
            fake = os.path.join(tmpdir, f"budget-buddy-2026-01-{i+1:02d}.xlsx")
            with open(fake, "wb") as f:
                f.write(b"fake")

        generate_snapshot(db_session, output_dir=tmpdir)
        files = [f for f in os.listdir(tmpdir) if f.endswith(".xlsx")]
        assert len(files) == 12
