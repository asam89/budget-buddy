"""Export API: CSV/XLSX for transactions, reports, and full workbook."""

import csv
import io
from datetime import datetime, date
from typing import Optional
from collections import defaultdict
from calendar import monthrange

from fastapi import APIRouter, Depends, Query, Response
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import or_, select

from openpyxl import Workbook
from openpyxl.styles import Font, numbers, Alignment, PatternFill
from openpyxl.utils import get_column_letter

from app.database import get_db
from app.models import (
    Account, Category, Entity, Transaction, TransactionSplit, User,
)
from app.utils.auth import get_current_user

router = APIRouter(prefix="/api/export", tags=["export"])


# ---- helpers ----

def _fmt(x: float) -> float:
    return round(x, 2)


def _get_filtered_txns(
    db: Session,
    account_id: Optional[int] = None,
    category_id: Optional[int] = None,
    entity_id: Optional[int] = None,
    txn_type: Optional[str] = None,
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    q: Optional[str] = None,
) -> list[Transaction]:
    query = db.query(Transaction).filter(Transaction.review_status == "confirmed")
    if account_id:
        query = query.filter(Transaction.account_id == account_id)
    if category_id:
        query = query.filter(Transaction.category_id == category_id)
    if entity_id:
        split_txn_ids = select(TransactionSplit.transaction_id).where(
            TransactionSplit.entity_id == entity_id
        )
        query = query.filter(
            or_(Transaction.entity_id == entity_id, Transaction.id.in_(split_txn_ids))
        )
    if txn_type:
        query = query.filter(Transaction.txn_type == txn_type)
    if start_date:
        query = query.filter(Transaction.date >= start_date)
    if end_date:
        query = query.filter(Transaction.date <= end_date)
    if q:
        pattern = f"%{q}%"
        query = query.filter(
            or_(Transaction.name.ilike(pattern), Transaction.merchant_name.ilike(pattern))
        )
    return query.order_by(Transaction.date.desc()).all()


def _txn_row(txn: Transaction, cat_map: dict, ent_map: dict, acct_map: dict) -> list:
    return [
        str(txn.date) if txn.date else "",
        txn.name or "",
        txn.merchant_name or "",
        txn.amount,
        txn.currency or "CAD",
        acct_map.get(txn.account_id, ""),
        cat_map.get(txn.category_id, ""),
        ent_map.get(txn.entity_id, ""),
        txn.txn_type or "",
        txn.review_status or "",
        txn.notes or "",
    ]


TXN_HEADERS = [
    "Date", "Name", "Merchant", "Amount", "Currency",
    "Account", "Category", "Entity", "Type", "Status", "Notes",
]


def _build_lookup_maps(db: Session):
    cat_map = {c.id: c.name for c in db.query(Category).all()}
    ent_map = {e.id: e.name for e in db.query(Entity).all()}
    acct_map = {a.id: a.name for a in db.query(Account).all()}
    return cat_map, ent_map, acct_map


# ---- CSV export ----

@router.get("/transactions/csv")
def export_transactions_csv(
    account_id: Optional[int] = None,
    category_id: Optional[int] = None,
    entity_id: Optional[int] = None,
    txn_type: Optional[str] = None,
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    q: Optional[str] = None,
    db: Session = Depends(get_db),
    _user: User = Depends(get_current_user),
):
    txns = _get_filtered_txns(db, account_id, category_id, entity_id, txn_type, start_date, end_date, q)
    cat_map, ent_map, acct_map = _build_lookup_maps(db)

    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(TXN_HEADERS)
    for txn in txns:
        writer.writerow(_txn_row(txn, cat_map, ent_map, acct_map))

    buf.seek(0)
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename=transactions-{date.today()}.csv"},
    )


# ---- XLSX export (single sheet) ----

def _style_header_row(ws, col_count: int):
    header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    header_font = Font(bold=True, color="E5E7EB")
    for col in range(1, col_count + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
    ws.freeze_panes = "A2"


@router.get("/transactions/xlsx")
def export_transactions_xlsx(
    account_id: Optional[int] = None,
    category_id: Optional[int] = None,
    entity_id: Optional[int] = None,
    txn_type: Optional[str] = None,
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    q: Optional[str] = None,
    db: Session = Depends(get_db),
    _user: User = Depends(get_current_user),
):
    txns = _get_filtered_txns(db, account_id, category_id, entity_id, txn_type, start_date, end_date, q)
    cat_map, ent_map, acct_map = _build_lookup_maps(db)

    wb = Workbook()
    ws = wb.active
    ws.title = "Transactions"

    ws.append(TXN_HEADERS)
    _style_header_row(ws, len(TXN_HEADERS))

    for txn in txns:
        ws.append(_txn_row(txn, cat_map, ent_map, acct_map))

    # Currency format for Amount column (col 4)
    for row in range(2, ws.max_row + 1):
        ws.cell(row=row, column=4).number_format = '#,##0.00'

    # Auto-width
    for col in range(1, len(TXN_HEADERS) + 1):
        ws.column_dimensions[get_column_letter(col)].width = max(12, len(TXN_HEADERS[col - 1]) + 4)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=transactions-{date.today()}.xlsx"},
    )


# ---- Full workbook export ----

def _add_info_sheet(wb: Workbook, filters: dict):
    ws = wb.create_sheet("Info", 0)
    ws.append(["Budget Buddy Export"])
    ws.cell(row=1, column=1).font = Font(bold=True, size=14)
    ws.append([])
    ws.append(["Generated", datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S UTC")])
    ws.append(["Filters", str(filters) if filters else "None (full export)"])
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 60


def _add_transactions_sheet(wb: Workbook, db: Session):
    ws = wb.create_sheet("Transactions")
    cat_map, ent_map, acct_map = _build_lookup_maps(db)
    txns = db.query(Transaction).filter(
        Transaction.review_status == "confirmed"
    ).order_by(Transaction.date.desc()).all()

    ws.append(TXN_HEADERS)
    _style_header_row(ws, len(TXN_HEADERS))
    for txn in txns:
        ws.append(_txn_row(txn, cat_map, ent_map, acct_map))
    for row in range(2, ws.max_row + 1):
        ws.cell(row=row, column=4).number_format = '#,##0.00'
    for col in range(1, len(TXN_HEADERS) + 1):
        ws.column_dimensions[get_column_letter(col)].width = max(12, len(TXN_HEADERS[col - 1]) + 4)


def _add_entities_sheet(wb: Workbook, db: Session):
    ws = wb.create_sheet("Entities")
    headers = ["ID", "Name", "Type", "Is Default", "Is Active"]
    ws.append(headers)
    _style_header_row(ws, len(headers))
    for e in db.query(Entity).order_by(Entity.name).all():
        ws.append([e.id, e.name, e.entity_type, e.is_default, e.is_active])
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 16


def _add_categories_sheet(wb: Workbook, db: Session):
    ws = wb.create_sheet("Categories")
    headers = ["ID", "Name", "Parent ID", "Is System"]
    ws.append(headers)
    _style_header_row(ws, len(headers))
    for c in db.query(Category).order_by(Category.name).all():
        ws.append([c.id, c.name, c.parent_id, c.is_system])
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 16


def _add_accounts_sheet(wb: Workbook, db: Session):
    ws = wb.create_sheet("Accounts")
    headers = ["ID", "Name", "Type", "Currency", "Balance", "Is Active"]
    ws.append(headers)
    _style_header_row(ws, len(headers))
    for a in db.query(Account).order_by(Account.name).all():
        ws.append([a.id, a.name, a.account_type, a.currency, a.current_balance, a.is_active])
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["E"].width = 16
    for row in range(2, ws.max_row + 1):
        ws.cell(row=row, column=5).number_format = '#,##0.00'


def _month_range(start: date, end: date) -> list[str]:
    months = []
    cur = start.replace(day=1)
    while cur <= end:
        months.append(cur.strftime("%Y-%m"))
        if cur.month == 12:
            cur = cur.replace(year=cur.year + 1, month=1)
        else:
            cur = cur.replace(month=cur.month + 1)
    return months


def _effective_amount(txn: Transaction, entity_id: int) -> float:
    if txn.splits:
        return sum(s.amount for s in txn.splits if s.entity_id == entity_id)
    return txn.amount


def _add_pivot_sheet(wb: Workbook, db: Session, entity: Entity, months: int = 12):
    """Category × Month pivot sheet for one entity."""
    ws = wb.create_sheet(f"Pivot – {entity.name}"[:31])  # sheet name max 31 chars

    now = datetime.utcnow().date()
    start = now.replace(day=1)
    for _ in range(months - 1):
        if start.month == 1:
            start = start.replace(year=start.year - 1, month=12)
        else:
            start = start.replace(month=start.month - 1)

    month_labels = _month_range(start, now)
    cat_map = {c.id: c.name for c in db.query(Category).all()}

    # Get entity transactions
    split_txn_ids = select(TransactionSplit.transaction_id).where(
        TransactionSplit.entity_id == entity.id
    )
    txns = (
        db.query(Transaction)
        .filter(
            Transaction.date >= start,
            Transaction.date <= now,
            Transaction.review_status == "confirmed",
            or_(
                Transaction.entity_id == entity.id,
                Transaction.id.in_(split_txn_ids),
            ),
        )
        .all()
    )

    matrix: dict[str, dict[str, float]] = defaultdict(lambda: defaultdict(float))
    for txn in txns:
        ym = txn.date.strftime("%Y-%m")
        if ym not in month_labels:
            continue
        cat_name = cat_map.get(txn.category_id, "Uncategorized") if txn.category_id else "Uncategorized"
        amt = _effective_amount(txn, entity.id)
        matrix[cat_name][ym] += amt

    # Write header
    headers = ["Category"] + month_labels + ["Total", "Average"]
    ws.append(headers)
    _style_header_row(ws, len(headers))

    # Write rows
    for cat_name in sorted(matrix.keys()):
        row = [cat_name]
        values = []
        for m in month_labels:
            v = _fmt(matrix[cat_name].get(m, 0))
            values.append(v)
            row.append(v)
        total = _fmt(sum(values))
        avg = _fmt(total / len(month_labels)) if month_labels else 0
        row.extend([total, avg])
        ws.append(row)

    # Format currency columns
    for row in range(2, ws.max_row + 1):
        for col in range(2, len(headers) + 1):
            ws.cell(row=row, column=col).number_format = '#,##0.00'

    ws.column_dimensions["A"].width = 24
    for col in range(2, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 14


@router.get("/full")
def export_full_workbook(
    db: Session = Depends(get_db),
    _user: User = Depends(get_current_user),
):
    """Full XLSX workbook: Info, Transactions, Entities, Categories, Accounts, + Pivot per entity."""
    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    _add_info_sheet(wb, {})
    _add_transactions_sheet(wb, db)
    _add_entities_sheet(wb, db)
    _add_categories_sheet(wb, db)
    _add_accounts_sheet(wb, db)

    # Per-entity pivot sheets
    entities = db.query(Entity).filter(Entity.is_active == True).all()
    for entity in entities:
        _add_pivot_sheet(wb, db, entity)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    fname = f"budget-buddy-{date.today()}.xlsx"
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"},
    )


# ---- Snapshot task (callable from scheduler or CLI) ----

def generate_snapshot(db: Session, output_dir: str = "data/snapshots"):
    """Generate a full export XLSX and save to disk. Called by scheduler."""
    import os
    os.makedirs(output_dir, exist_ok=True)

    wb = Workbook()
    wb.remove(wb.active)

    _add_info_sheet(wb, {"type": "scheduled_snapshot"})
    _add_transactions_sheet(wb, db)
    _add_entities_sheet(wb, db)
    _add_categories_sheet(wb, db)
    _add_accounts_sheet(wb, db)

    entities = db.query(Entity).filter(Entity.is_active == True).all()
    for entity in entities:
        _add_pivot_sheet(wb, db, entity)

    fname = f"budget-buddy-{date.today()}.xlsx"
    path = os.path.join(output_dir, fname)
    wb.save(path)

    # Keep only the last 12 snapshots
    files = sorted(
        [f for f in os.listdir(output_dir) if f.endswith(".xlsx")],
        reverse=True,
    )
    for old_file in files[12:]:
        os.remove(os.path.join(output_dir, old_file))

    return path
